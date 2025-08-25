"""
새 학생 정보를 직접 추가하는 간단한 스크립트
- 엑셀 파일에 새 행을 직접 추가하는 방식
"""
import openpyxl
import os

def add_new_student(student_id, name, seat):
    """학생 정보를 Excel 파일에 직접 추가 (헤더 없는 구조 대응)"""
    excel_file = 'students.xlsx'
    
    # 엑셀 파일이 없는 경우, 새로 생성
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        # 첫 번째 데이터로 바로 추가 (헤더 없이)
        ws.cell(row=1, column=1).value = student_id
        ws.cell(row=1, column=2).value = name
        ws.cell(row=1, column=3).value = seat
        wb.save(excel_file)
        return True, f"학생 추가 성공: {student_id}, {name}, {seat}"
    else:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    
    # 현재 파일 구조 확인 (첫 번째 행이 헤더인지 데이터인지)
    first_row = [cell.value for cell in ws[1]]
    has_header = False
    
    # 첫 번째 행이 숫자(학번)로 시작하면 헤더 없음
    if first_row[0] and str(first_row[0]).isdigit():
        has_header = False
        start_row = 1
    else:
        has_header = True
        start_row = 2
    
    # 학번 중복 검사
    student_ids = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if row[0]:  # 첫 번째 열(학번)이 있는 경우만
            student_ids.append(str(row[0]))
    
    # 이미 존재하면 메시지 반환
    if student_id in student_ids:
        return False, f"학번 {student_id}는 이미 존재합니다."
    
    # 새로운 행에 학생 정보 추가 (현재 구조: 학번, 이름, 좌석번호)
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1).value = student_id
    ws.cell(row=next_row, column=2).value = name
    ws.cell(row=next_row, column=3).value = seat
    
    try:
        # 파일 저장
        wb.save(excel_file)
        return True, f"학생 추가 성공: {student_id}, {name}, {seat}"
    except Exception as e:
        return False, f"저장 중 오류: {str(e)}"
    
if __name__ == "__main__":
    # 테스트 코드
    success, message = add_new_student("10701", "한가람", "600")
    print(message)