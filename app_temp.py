"""
임시 파일: 학생 정보 일괄 수정 기능 구현
- 이름과 좌석번호를 함께 수정할 수 있도록 업데이트
- 이 파일의 코드 내용을 app.py의 해당 부분에 반영
"""

# API 엔드포인트 코드를 다음과 같이 변경
@app.route('/api/bulk_update_seats', methods=['POST'])
def api_bulk_update_seats():
    """학생 좌석번호 및 이름 일괄 업데이트 API (관리자 전용)"""
    if not session.get('admin'):
        return jsonify({"error": "관리자 권한이 필요합니다."}), 403
    
    data = request.json
    if not data or 'changes' not in data or not data['changes']:
        return jsonify({"error": "유효한 데이터가 없습니다."}), 400
    
    try:
        # Excel 파일 업데이트
        import pandas as pd
        from openpyxl import load_workbook
        
        excel_path = 'students.xlsx'
        
        # pandas로 열 이름 확인
        df_preview = pd.read_excel(excel_path, nrows=1)
        column_names = df_preview.columns.tolist()
        
        # 학번, 이름, 좌석번호 열 확인
        id_column = [col for col in column_names if '학번' in col or 'ID' in col.upper()][0]
        seat_column = [col for col in column_names if '좌석' in col or 'SEAT' in col.upper()][0]
        name_columns = [col for col in column_names if '이름' in col or 'NAME' in col.upper()]
        name_column = name_columns[0] if name_columns else None
        
        # openpyxl로 직접 셀 수정
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 학번, 이름, 좌석번호 열의 인덱스 찾기
        col_indices = {}
        for idx, col in enumerate(ws[1]):
            if col.value == id_column:
                col_indices['id'] = idx + 1  # 1-based index
            elif col.value == seat_column:
                col_indices['seat'] = idx + 1  # 1-based index
            elif name_column and col.value == name_column:
                col_indices['name'] = idx + 1  # 1-based index
        
        # 변경 사항 추적
        seat_changes_count = 0
        name_changes_count = 0
        not_found_count = 0
        
        # 학번별 새 좌석번호/이름 매핑 생성
        changes_map = {}
        for item in data['changes']:
            student_id = item['student_id']
            changes_map[student_id] = {
                'new_seat': item.get('new_seat')
            }
            # 새 이름이 있는 경우만 추가
            if 'new_name' in item and item['new_name']:
                changes_map[student_id]['new_name'] = item['new_name']
        
        # 엑셀 파일 수정
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # 2행부터 시작 (헤더 제외)
            cell_id = str(row[col_indices['id'] - 1].value).strip()  # 0-based index
            if cell_id in changes_map:
                change_item = changes_map[cell_id]
                
                # 좌석번호 업데이트
                if 'new_seat' in change_item and change_item['new_seat']:
                    ws.cell(row=row_idx, column=col_indices['seat']).value = change_item['new_seat']
                    seat_changes_count += 1
                
                # 이름 업데이트 (이름 칼럼이 존재하고 새 이름이 제공된 경우)
                if 'name' in col_indices and 'new_name' in change_item and change_item['new_name']:
                    ws.cell(row=row_idx, column=col_indices['name']).value = change_item['new_name']
                    name_changes_count += 1
                
                # 처리한 항목 제거
                del changes_map[cell_id]
        
        # 미처리된 학번 수 계산
        not_found_count = len(changes_map)
        
        # 변경 사항 저장
        wb.save(excel_path)
        
        # 학생 데이터 캐시 초기화
        global _student_data_cache, _student_data_timestamp
        _student_data_cache = None
        _student_data_timestamp = None
        
        return jsonify({
            "success": True,
            "message": f"성공적으로 업데이트되었습니다: 좌석번호 {seat_changes_count}개, 이름 {name_changes_count}개. {not_found_count}개의 학번은 찾을 수 없습니다."
        })
    except Exception as e:
        return jsonify({"error": f"학생 정보 일괄 업데이트 중 오류가 발생했습니다: {str(e)}"}), 500