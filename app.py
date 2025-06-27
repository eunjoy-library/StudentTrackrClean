import os
import json
import logging
import time
from datetime import datetime, timedelta
from collections import Counter

import pytz
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory
from dotenv import load_dotenv

# Firebase 관련 라이브러리
import firebase_admin
from firebase_admin import credentials, firestore

# 환경 변수 로드
load_dotenv(override=True)

# 한국 시간대 설정
KST = pytz.timezone('Asia/Seoul')

# Firebase 초기화
db = None
try:
    FIREBASE_CREDENTIALS_JSON = os.environ.get("FIREBASE_CREDENTIALS_JSON")
    if not FIREBASE_CREDENTIALS_JSON:
        raise ValueError("FIREBASE_CREDENTIALS_JSON 환경변수가 없습니다.")
    
    # JSON 문자열을 딕셔너리로 변환
    firebase_config = json.loads(FIREBASE_CREDENTIALS_JSON)
    
    # Firebase 인증 정보 생성
    cred = credentials.Certificate(firebase_config)
    
    # Firebase 앱 초기화 (이미 초기화되었는지 확인)
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    
    # Firestore 클라이언트 설정
    db = firestore.client()
    print("Firebase 초기화 성공")
except Exception as e:
    print(f"Firebase 초기화 오류: {e}")

# 로깅 설정    
logging.basicConfig(level=logging.DEBUG, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Flask 앱 초기화
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "your-secret-key")

# 캐싱을 위한 전역 변수
_student_data_cache = None
_last_student_data_load_time = None

# ================== [UTILITY 함수] ==================
def get_current_period():
    """
    현재 시간 기준으로 교시를 결정하는 함수
    1교시: 7:50 - 9:15
    2교시: 9:15 - 10:40
    3교시: 10:40 - 12:05
    4교시: 12:05 - 12:30
    5교시: 12:30 - 14:25
    6교시: 14:25 - 15:50
    
    Returns:
        int: 1~10 교시, -1 (시간 외), 0 (4교시)
    """
    now = datetime.now(KST)
    current_time = now.time()
    
    # 요일이 토요일(5) 또는 일요일(6)인 경우 시간 외로 처리
    if now.weekday() > 4:  # 토요일(5), 일요일(6)
        return -1
    
    # 교시 시간대 정의 (datetime 모듈에서 직접 time 사용)
    from datetime import time
    periods = [
        (time(7, 50), time(9, 15), 1),   # 1교시
        (time(9, 15), time(10, 40), 2),  # 2교시
        (time(10, 40), time(12, 5), 3),  # 3교시
        (time(12, 5), time(12, 30), 0),  # 4교시 (도서실 이용 불가 시간)
        (time(12, 30), time(14, 25), 5), # 5교시
        (time(14, 25), time(15, 50), 6)  # 6교시
    ]
    
    # 현재 시간이 어느 교시에 해당하는지 확인
    for start, end, period in periods:
        if start <= current_time < end:
            return period
    
    # 어느 교시에도 해당하지 않으면 시간 외로 처리
    return -1

def load_student_data(force_reload=False):
    """
    Load student data from Excel file with caching
    Returns a dictionary with student_id as key and (name, seat) as value
    
    Args:
        force_reload: 강제로 새로 로딩할지 여부 (기본값: False)
    """
    global _student_data_cache, _last_student_data_load_time
    now = datetime.now()
    
    # 캐시 완전 제거 (문제 해결을 위해)
    if force_reload:
        _student_data_cache = None
        logging.debug("학생 데이터 캐시 강제 초기화")
    
    # 캐시가 없거나 5초 이상 지난 경우 새로 로드 (개선된 방식)
    if (_student_data_cache is None or 
            _last_student_data_load_time is None or 
            (now - _last_student_data_load_time).seconds > 5):
        logging.debug("학생 데이터 새로 로딩")
        try:
            # openpyxl을 직접 사용하여 엑셀 파일 읽기 (더 안정적인 방식)
            try:
                import openpyxl
                wb = openpyxl.load_workbook('students.xlsx')
                ws = wb.active
                
                student_data = {}
                headers = []
                seat_column = 2  # 기본 열 인덱스
                
                # 헤더 행 읽기 (첫 번째 행)
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # 좌석번호 또는 공강좌석번호 열 찾기
                for i, header in enumerate(headers):
                    if header in ['좌석번호', '공강좌석번호']:
                        seat_column = i
                        break
                
                # 학번과 이름 열 인덱스 찾기
                student_id_column = -1
                name_column = -1
                for i, header in enumerate(headers):
                    if header == '학번':
                        student_id_column = i
                    elif header == '이름':
                        name_column = i
                
                if student_id_column == -1 or name_column == -1:
                    logging.error(f"헤더 구조 문제: 학번 또는 이름 열을 찾을 수 없음. 헤더: {headers}")
                    return {}
                
                # 학생 데이터 읽기 (2번째 행부터)
                for row in ws.iter_rows(min_row=2):
                    student_id = row[student_id_column].value
                    name = row[name_column].value
                    seat = row[seat_column].value if len(row) > seat_column else ''
                    
                    # 값이 모두 있는 경우만 추가
                    if student_id and name:
                        student_id = str(student_id).strip()
                        name = str(name).strip()
                        seat = str(seat).strip() if seat else ''
                        student_data[student_id] = (name, seat)
                
                logging.debug(f"로드된 학생 수: {len(student_data)}")
                
            except Exception as excel_error:
                logging.error(f"엑셀 파일 직접 읽기 실패: {excel_error}")
                
                # pandas 방식으로 시도 (대체 방식)
                try:
                    df = pd.read_excel('students.xlsx', dtype={'학번': str})
                    # 컬럼명 확인 (좌석번호 또는 공강좌석번호)
                    seat_column = '공강좌석번호' if '공강좌석번호' in df.columns else '좌석번호'
                    
                    student_data = {}
                    for _, row in df.iterrows():
                        # 학번이 있는 경우만 처리
                        if pd.notna(row['학번']) and pd.notna(row['이름']):
                            student_id = str(row['학번']).strip()
                            name = str(row['이름']).strip()
                            # 좌석번호 가져오기 (없으면 빈 문자열)
                            seat = str(row.get(seat_column, '')) if pd.notna(row.get(seat_column, '')) else ''
                            student_data[student_id] = (name, seat)
                    
                    logging.debug(f"pandas로 로드된 학생 수: {len(student_data)}")
                    
                except Exception as pandas_error:
                    logging.error(f"pandas 읽기 실패: {pandas_error}")
                    # 임시 학생 데이터 제공 (테스트용)
                    student_data = {
                        "10307": ("박지호", "387"),
                        "20101": ("강지훈", "331"),
                        "30107": ("김리나", "175"),
                        "30207": ("김유담", "281"),
                        "20240101": ("홍길동", "A1"),
                        "10701": ("한가람", "600"),  # 테스트 학생 추가
                        # 추가 데이터...
                    }
            
            _student_data_cache = student_data
            _last_student_data_load_time = now
            
            # 디버깅을 위해 캐시에 10701 학번이 있는지 확인
            if '10701' in _student_data_cache:
                logging.debug(f"10701 학번 데이터 있음: {_student_data_cache.get('10701')}")
            else:
                logging.debug("10701 학번 데이터 없음")
                
            return student_data
            
        except Exception as e:
            logging.error(f"학생 데이터 로딩 중 오류: {e}")
            return {}
    else:
        logging.debug("학생 데이터 캐시 사용")
        return _student_data_cache

def save_attendance(student_id, name, seat, period_text, admin_override=False):
    """
    출석 기록을 학번별로 조직화하여 저장 (한국 시간 기준)
    구조: students/{student_id}/attendance/{date}
    - 읽기 횟수 대폭 감소
    - admin_override: 관리자 권한으로 중복 출석 허용 여부
    """
    try:
        # Firebase 연결 확인
        if not db:
            flash("Firebase 설정이 완료되지 않았습니다.", "danger")
            return False
        
        # 현재 시간으로 출석 기록 생성
        now_kst = datetime.now(KST)
        date_str = now_kst.strftime('%Y-%m-%d')
        datetime_str = now_kst.strftime('%Y-%m-%d %H:%M:%S')
        
        # 관리자 모드가 아닌 경우 중복 체크
        if not admin_override:
            # 현재 주 범위 계산 (일~토)
            days_to_sunday = now_kst.weekday() + 1
            if days_to_sunday == 7:
                days_to_sunday = 0
                
            sunday = now_kst - timedelta(days=days_to_sunday)
            sunday_str = sunday.strftime('%Y-%m-%d')
            
            saturday = sunday + timedelta(days=6)
            saturday_str = saturday.strftime('%Y-%m-%d')
            
            # 새 구조에서 학번별 출석 기록만 확인 (매우 효율적)
            student_ref = db.collection('attendance').document(student_id).collection('records')
            
            # 이번 주 범위의 출석 기록만 조회
            week_records = student_ref.where('date_only', '>=', sunday_str).where('date_only', '<=', saturday_str).get()
            
            if len(week_records) > 0:
                # 이미 출석한 날짜 찾기
                attendance_date = week_records[0].to_dict().get('date_only', '')
                flash(f'이미 이번 주에 출석 기록이 있습니다. (출석일: {attendance_date})', 'warning')
                return False
        
        # 새로운 이중 구조로 출석 기록 저장
        attendance_data = {
            'student_id': student_id,
            'name': name,
            'seat': seat,
            'period': period_text,
            'date': datetime_str,
            'date_only': date_str,
            'timestamp': firestore.SERVER_TIMESTAMP
        }
        
        # 1. 학생별 출석 기록 저장 (학생 중복 체크용)
        try:
            student_attendance_ref = db.collection('attendance').document(student_id).collection('records').document(date_str)
            student_attendance_ref.set(attendance_data)
            logging.info(f"attendance/{student_id}/records/{date_str} 저장 성공")
        except Exception as e:
            logging.error(f"attendance 저장 실패: {e}")
            return False
        
        # 2. 관리자용 날짜+교시별 출석 기록 저장 (관리자 현황 파악용)
        try:
            date_period_key = f"{date_str}_{period_text}"
            admin_ref = db.collection('admin').document(date_period_key).collection('students').document(student_id)
            admin_ref.set({
                'student_id': student_id,
                'name': name,
                'seat': seat,
                'date': datetime_str,
                'date_only': date_str,
                'period': period_text,
                'timestamp': firestore.SERVER_TIMESTAMP
            })
            logging.info(f"admin/{date_period_key}/students/{student_id} 저장 성공")
        except Exception as e:
            logging.error(f"admin 저장 실패: {e}")
            return False
        
        # CSV 파일에도 즉시 저장 (교시별 출석 현황 즉시 반영용)
        try:
            csv_path = 'attendance.csv'
            
            # 기존 CSV 파일이 있으면 읽어서 추가
            if os.path.exists(csv_path):
                # 기존 CSV 파일에 새 행 추가 (한글 헤더 사용)
                with open(csv_path, 'a', encoding='utf-8', newline='') as f:
                    f.write(f'"{datetime_str}","{period_text}","{student_id}","{name}","{seat}"\n')
                logging.info(f"CSV 파일에 출석 기록 추가: {name}")
            else:
                # 새 CSV 파일 생성 (한글 헤더 포함)
                with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                    f.write('출석일,교시,학번,이름,공강좌석번호\n')
                    f.write(f'"{datetime_str}","{period_text}","{student_id}","{name}","{seat}"\n')
                logging.info(f"새 CSV 파일 생성 및 출석 기록 추가: {name}")
                
        except Exception as csv_error:
            logging.warning(f"CSV 저장 실패 (Firebase 저장은 성공): {csv_error}")
        
        # 저장 완료
        logging.info(f"출석 기록 저장 완료: {student_id} ({name}) - {period_text}")
        
        logging.info(f"학생 {student_id}({name})의 출석이 성공적으로 등록되었습니다. 좌석: {seat}, 날짜: {date_str}")
        
        # 캐시 초기화
        global attendance_status_cache
        attendance_status_cache.clear()
        
        return True
        
    except Exception as e:
        logging.error(f"출석 저장 중 오류 발생: {e}")
        return False

def load_attendance():
    """
    출석 기록을 CSV 파일과 Firebase에서 로드하는 통합 방식
    Firebase가 실패하는 경우 CSV를 백업으로 사용
    """
    try:
        attendance_records = []
        
        # 1. CSV 파일에서 출석 기록 로드 (백업 시스템)
        csv_path = 'attendance.csv'
        if os.path.exists(csv_path):
            try:
                df = pd.read_csv(csv_path, encoding='utf-8')
                logging.info(f"CSV에서 {len(df)}개 출석 기록 로드")
                
                for _, row in df.iterrows():
                    # 한글 헤더로 접근
                    student_id = str(row.get('학번', ''))
                    name = str(row.get('이름', ''))
                    seat = str(row.get('공강좌석번호', ''))
                    period = str(row.get('교시', ''))
                    date = str(row.get('출석일', ''))
                    
                    record = {
                        'id': f"csv_{student_id}_{date}",
                        'student_id': student_id,
                        'name': name,
                        'seat': seat,
                        'period': period,
                        'date': date,
                        'date_only': date[:10] if pd.notna(row.get('출석일')) else '',
                        'source': 'csv'
                    }
                    attendance_records.append(record)
                    
            except Exception as csv_error:
                logging.error(f"CSV 로드 실패: {csv_error}")
        
        # 2. Firebase에서 추가 데이터 로드 (가능한 경우)
        if db:
            try:
                # admin 컬렉션에서 최신 데이터 확인
                admin_docs = list(db.collection('admin').limit(50).get())
                logging.info(f"Firebase admin 컬렉션: {len(admin_docs)}개 문서")
                
                # admin 컬렉션에서 데이터 추가
                for admin_doc in admin_docs:
                    date_period = admin_doc.id
                    students = list(admin_doc.reference.collection('students').get())
                    
                    for student_doc in students:
                        data = student_doc.to_dict()
                        if data:
                            # CSV에 동일한 기록이 없는 경우만 추가
                            record_id = f"{data.get('student_id')}_{data.get('date_only')}"
                            existing = any(r['id'].endswith(record_id) for r in attendance_records)
                            
                            if not existing:
                                data['id'] = f"firebase_{date_period}_{student_doc.id}"
                                data['source'] = 'firebase'
                                attendance_records.append(data)
                                logging.debug(f"Firebase admin에서 추가: {data.get('name')}")
                
                # attendance CSV 파일에서 누락된 최신 기록을 Firebase admin 컬렉션에서 직접 추가
                try:
                    # 오늘 날짜의 모든 교시별 기록을 admin 컬렉션에서 직접 조회
                    today = datetime.now(KST).strftime('%Y-%m-%d')
                    periods = ['1교시', '2교시', '3교시', '4교시', '5교시', '6교시', '7교시', '8교시', '9교시', '10교시', '시간 외']
                    
                    for period in periods:
                        date_period_key = f"{today}_{period}"
                        try:
                            admin_doc = db.collection('admin').document(date_period_key).get()
                            if admin_doc.exists:
                                # 해당 교시의 모든 학생 조회
                                students_collection = admin_doc.reference.collection('students')
                                students_docs = list(students_collection.get())
                                
                                for student_doc in students_docs:
                                    data = student_doc.to_dict()
                                    if data:
                                        # CSV에 동일한 기록이 없는 경우만 추가
                                        record_id = f"{data.get('student_id')}_{today}"
                                        existing = any(record_id in r.get('id', '') for r in attendance_records)
                                        
                                        if not existing:
                                            data['id'] = f"firebase_admin_{date_period_key}_{student_doc.id}"
                                            data['source'] = 'firebase_admin'
                                            attendance_records.append(data)
                                            logging.info(f"Firebase admin에서 {period} 추가: {data.get('name')}")
                        
                        except Exception as period_error:
                            logging.debug(f"Firebase {date_period_key} 조회 실패: {period_error}")
                            
                except Exception as admin_error:
                    logging.warning(f"Firebase admin 세부 조회 실패: {admin_error}")
                                
            except Exception as firebase_error:
                logging.warning(f"Firebase 로드 실패 (CSV 백업 사용): {firebase_error}")
        
        # 최종 정렬 및 반환
        final_records = sorted(attendance_records, key=lambda x: x.get('date', ''), reverse=True)
        logging.info(f"최종 출석 기록: {len(final_records)}개 (CSV: {sum(1 for r in final_records if r.get('source') == 'csv')}, Firebase: {sum(1 for r in final_records if r.get('source') == 'firebase')})")
        
        return final_records
        
    except Exception as e:
        logging.error(f"출석 기록 로딩 중 치명적 오류: {e}")
        return []

# 출석 상태 캐싱을 위한 딕셔너리와 캐시 만료 시간 (초)
attendance_status_cache = {}
CACHE_EXPIRY = 5  # 5초로 대폭 감소 (중복 출석 방지 강화)
student_data_cache = {}  # 학생 정보 캐시
STUDENT_CACHE_EXPIRY = 600  # 학생 정보는 10분 동안 캐시 (성능 향상)

# 중복 출석 방지를 위한 잠금 메커니즘 (동시 요청 처리용)
attendance_locks = {}

# 이번 주 날짜 범위 정보 캐싱 (매번 계산하지 않도록)
current_week_info = {
    'last_updated': 0,
    'sunday_str': '',
    'saturday_str': ''
}
WEEK_INFO_EXPIRY = 3600  # 주 정보는 1시간마다 갱신

def get_current_week_range():
    """현재 주의 시작(일요일)과 끝(토요일) 계산 - 캐싱 적용"""
    global current_week_info
    
    now = time.time()
    if now - current_week_info['last_updated'] < WEEK_INFO_EXPIRY and current_week_info['sunday_str']:
        return current_week_info['sunday_str'], current_week_info['saturday_str']
    
    # 현재 날짜 기준으로 이번 주의 월요일 찾기
    now_date = datetime.now(KST)
    weekday = now_date.weekday()  # 0=월요일, 1=화요일, ..., 6=일요일
    days_since_monday = weekday
    monday = now_date - timedelta(days=days_since_monday)
    monday = monday.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # 이번 주 토요일 끝 시간 (일요일부터 토요일까지)
    sunday = monday - timedelta(days=1)  # 일요일은 월요일 하루 전
    sunday = sunday.replace(hour=0, minute=0, second=0, microsecond=0)
    
    saturday = monday + timedelta(days=5)  # 토요일은 월요일부터 5일 후
    saturday = saturday.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # 텍스트 형식으로 변환
    sunday_str = sunday.strftime('%Y-%m-%d')
    saturday_str = saturday.strftime('%Y-%m-%d')
    
    # 결과 캐싱
    current_week_info = {
        'last_updated': now,
        'sunday_str': sunday_str,
        'saturday_str': saturday_str
    }
    
    return sunday_str, saturday_str

def check_weekly_attendance_limit(student_id):
    """
    학번별 구조에서 학생의 이번 주 출석 확인 (초고속)
    구조: students/{student_id}/attendance/{date}
    - Firebase 읽기 횟수 최소화 (해당 학생만 조회)
    
    Returns:
        (exceeded, count, recent_dates): 
        - exceeded: 주 1회 초과 여부 (True/False)
        - count: 이번 주 출석 횟수
        - recent_dates: 최근 출석 날짜 목록
    """
    global attendance_status_cache
    
    # 캐시 확인
    cache_key = f"weekly_limit_{student_id}"
    if cache_key in attendance_status_cache:
        cache_entry = attendance_status_cache[cache_key]
        if time.time() - cache_entry['timestamp'] < CACHE_EXPIRY:
            logging.debug(f"학생 {student_id}의 출석 상태 캐시 사용")
            return cache_entry['exceeded'], cache_entry['count'], cache_entry['recent_dates']
    
    try:
        if not db:
            logging.error("Firebase DB 연결이 설정되지 않았습니다.")
            return False, 0, []
        
        # 현재 주 범위 계산
        now = datetime.now(KST)
        days_to_sunday = now.weekday() + 1
        if days_to_sunday == 7:
            days_to_sunday = 0
            
        sunday = now - timedelta(days=days_to_sunday)
        sunday_str = sunday.strftime('%Y-%m-%d')
        
        saturday = sunday + timedelta(days=6)
        saturday_str = saturday.strftime('%Y-%m-%d')
        
        logging.debug(f"학생 {student_id}의 이번 주({sunday_str} ~ {saturday_str}) 출석 기록 확인 중")
        
        # 새 구조에서 학번별 출석 기록만 조회 (매우 효율적)
        student_ref = db.collection('attendance').document(student_id).collection('records')
        week_records = student_ref.where('date_only', '>=', sunday_str).where('date_only', '<=', saturday_str).get()
        
        # 결과 처리
        count = len(week_records)
        recent_dates = []
        
        for record in week_records:
            data = record.to_dict()
            date_only = data.get('date_only', '')
            recent_dates.append(date_only)
        
        # 결과 정리
        recent_dates = sorted(recent_dates, reverse=True)
        exceeded = count >= 1  # 1회 이상 출석했으면 제한
        
        # 캐시 저장
        attendance_status_cache[cache_key] = {
            'exceeded': exceeded,
            'count': count,
            'recent_dates': recent_dates,
            'timestamp': time.time()
        }
        
        logging.debug(f"학생 {student_id}의 이번 주 출석 횟수: {count}, 초과 여부: {exceeded}")
        return exceeded, count, recent_dates
        
    except Exception as e:
        logging.error(f"주간 출석 제한 확인 중 오류: {e}")
        return False, 0, []

# ================== [ROUTES] ==================

@app.route('/api/check_attendance', methods=['GET', 'POST'])
def api_check_attendance():
    """
    학생 ID로 해당 주에 출석 기록이 있는지 직접 확인하는 API
    - 정확한 기록 확인을 위해 캐시 없이 DB에서 직접 조회
    - 중복 출석 방지를 위한 실시간 검증
    - 경고받은 학생 출석 제한 기능 추가
    """
    student_id = request.args.get('student_id')
    
    # POST 파라미터에서도 확인
    if not student_id and request.method == 'POST':
        student_id = request.form.get('student_id')
    
    # 타임스탬프 확인 (캐시 방지)
    timestamp = request.args.get('t')
    
    if not student_id:
        return jsonify({'error': '학번이 필요합니다.', 'has_attendance': False})
    
    # 경고받은 학생인지 확인 (추가)
    try:
        # 'warnings' 컬렉션에서 해당 학생 ID에 대한 활성화된 경고 확인
        if db:
            warnings_query = db.collection('warnings').where('student_id', '==', student_id).where('active', '==', True)
            warning_docs = warnings_query.get()
            
            # 활성화된 경고가 있는지 확인
            for doc in warning_docs:
                warning_data = doc.to_dict()
                # 경고 상태인 경우 출석 제한
                return jsonify({
                    'warning': True,
                    'has_attendance': False,
                    'message': '경고 상태로 인해 출석이 제한되었습니다. 관리자에게 문의하세요.',
                    'warning_reason': warning_data.get('reason', '경고 상태')
                })
    except Exception as e:
        logging.error(f"경고 확인 오류: {e}")
        # 경고 확인 실패 시에도 계속 진행하여 출석은 확인
    
    try:
        # 현재 주의 범위 계산 (일요일부터 토요일까지)
        now = datetime.now(KST)
        
        # 이번 주 일요일(주 시작) 계산
        days_to_sunday = now.weekday() + 1  # 월=0, 일=6이므로 역으로 계산
        if days_to_sunday == 7:  # 일요일인 경우
            days_to_sunday = 0
            
        sunday = now - timedelta(days=days_to_sunday)
        sunday = sunday.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # 토요일(주 마지막) 계산
        days_to_saturday = (5 - now.weekday()) % 7  # 토요일까지 남은 일수
        saturday = now + timedelta(days=days_to_saturday)
        saturday = saturday.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        sunday_str = sunday.strftime('%Y-%m-%d')
        saturday_str = saturday.strftime('%Y-%m-%d')
        
        # Firebase 직접 쿼리
        if not db:
            logging.error("Firebase DB 연결이 설정되지 않았습니다.")
            return jsonify({'error': 'Firebase 연결 오류', 'has_attendance': False})
            
        try:
            # 학생 ID로 필터링하여 쿼리 실행 (인덱스 오류 없는 단순 쿼리)
            records = db.collection('attendances').where('student_id', '==', student_id).get()
            
            # 이번 주에 해당하는 기록만 필터링 (직접 확인)
            has_attendance = False
            attendance_date = ""
            recent_dates = []
            
            for record in records:
                data = record.to_dict()
                date_only = data.get('date_only', '')
                
                # 이번 주 날짜 범위에 있는지 확인 
                if sunday_str <= date_only <= saturday_str:
                    has_attendance = True
                    attendance_date = date_only
                    recent_dates.append(date_only)
            
            # 결과 정리 (캐시 사용 안함)
            recent_dates = sorted(recent_dates, reverse=True)  # 최신 날짜 먼저
            
            # 한국어 요일 추가
            formatted_date = ""
            if attendance_date:
                try:
                    # yyyy-mm-dd 형식의 날짜 문자열에서 datetime 객체로 변환
                    date_obj = datetime.strptime(attendance_date, '%Y-%m-%d')
                    # 한국어 요일
                    weekdays = ['월', '화', '수', '목', '금', '토', '일']
                    weekday_kr = weekdays[date_obj.weekday()]
                    # 날짜 형식: 5월 18일 (토)
                    formatted_date = f"{date_obj.month}월 {date_obj.day}일 ({weekday_kr})"
                except Exception as e:
                    logging.error(f"날짜 변환 중 오류: {e}")
                    formatted_date = attendance_date
            
            logging.info(f"학생 {student_id}의 이번 주 출석 상태: {has_attendance}, 출석일: {recent_dates}")
            
            return jsonify({
                'has_attendance': has_attendance,
                'attendance_date': attendance_date,
                'formatted_date': formatted_date,
                'cached': False,
                'timestamp': str(datetime.now(KST))
            })
            
        except Exception as db_error:
            logging.error(f"Firebase 쿼리 실행 중 오류: {db_error}")
            raise
        
    except Exception as e:
        logging.error(f"출석 확인 API 오류: {e}")
        return jsonify({'error': str(e), 'has_attendance': False})

@app.route('/check_attendance_status')
def check_attendance_status():
    """학생의 출석 상태를 확인하는 API (주간 출석 제한용)"""
    student_id = request.args.get('student_id')
    if not student_id:
        return jsonify({'error': '학번이 필요합니다.', 'already_attended': False})
    
    try:
        # 주간 출석 상태 확인
        exceeded, count, recent_dates = check_weekly_attendance_limit(student_id)
        
        # 최근 출석일 포맷팅
        last_attendance_date = ""
        if recent_dates:
            last_attendance_date = recent_dates[0]  # 가장 최근 출석일
        
        # 일주일에 두 번 오는 학생들을 위한 특별 처리
        twice_weekly_students = ['30530', '30606', '30607', '30608', '30609', '30610']  # 예시 학번들
        is_twice_weekly = student_id in twice_weekly_students
        
        # 두 번 출석 가능한 학생의 경우 처리 방식 변경
        if is_twice_weekly:
            if count == 0:
                # 첫 번째 출석
                return jsonify({
                    'already_attended': False,
                    'attendance_count': count,
                    'last_attendance_date': last_attendance_date,
                    'is_twice_weekly': True,
                    'show_twice_weekly_popup': False
                })
            elif count == 1:
                # 두 번째 출석 - 특별 팝업 표시
                return jsonify({
                    'already_attended': False,  # 아직 출석 가능
                    'attendance_count': count,
                    'last_attendance_date': last_attendance_date,
                    'is_twice_weekly': True,
                    'show_twice_weekly_popup': True
                })
            else:
                # 이미 두 번 출석함
                return jsonify({
                    'already_attended': True,
                    'attendance_count': count,
                    'last_attendance_date': last_attendance_date,
                    'is_twice_weekly': True,
                    'show_twice_weekly_popup': False
                })
        else:
            # 일반 학생 (주 1회만) - 이미 출석한 경우 즉시 차단
            return jsonify({
                'already_attended': exceeded,  # 1회 출석 후 True가 됨
                'attendance_count': count,
                'last_attendance_date': last_attendance_date,
                'is_twice_weekly': False,
                'show_twice_weekly_popup': False
            })
    except Exception as e:
        logging.error(f"출석 상태 확인 중 오류: {e}")
        return jsonify({'error': str(e), 'already_attended': False})

@app.route('/')
def index():
    """관리자 로그인 여부에 따라 적절한 페이지로 리다이렉트"""
    if session.get('admin'):
        return redirect(url_for('by_period'))  # 관리자면 교시별 보기로
    else:
        return redirect(url_for('attendance'))  # 일반 사용자면 출석 페이지로

@app.route('/favicon.ico')
def favicon():
    """Serve favicon.ico"""
    return send_from_directory(os.path.join(app.root_path, 'attached_assets'), 'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    """Main attendance page and form submission handler"""
    # 현재 시간과 교시 정보 계산
    now = datetime.now(KST)
    current_period = get_current_period()
    weekday_korean = ['월', '화', '수', '목', '금', '토', '일'][now.weekday()]
    
    # 교시 텍스트 생성
    if current_period == -1:
        period_text = "시간 외"
    elif current_period == 0:
        period_text = "4교시 (도서실 이용 불가)"
    else:
        period_text = f"{current_period}교시"
    
    # POST 요청 처리 (출석 폼 제출)
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        name = request.form.get('name', '').strip()
        seat = request.form.get('seat', '').strip()
        
        # 관리자 접근용 학번 체크 (환경 변수 사용으로 보안 강화)
        admin_access_id = os.environ.get('ADMIN_ACCESS_ID', 'ac97d429fb3e')
        if student_id == admin_access_id:
            # 관리자 페이지로 리다이렉트
            return redirect(url_for('admin_login'))
        
        if not student_id:
            flash('학번을 입력해주세요.', 'danger')
            return redirect(url_for('attendance'))
        
        # 주간 출석 제한 확인 (1회만 허용)
        # 출석 API로 직접 확인 (자바스크립트에서 이미 확인했지만 더블체크)
        try:
            # 직접 함수 호출로 변경 (API 호출 방식 문제 수정)
            exceeded, count, recent_dates = check_weekly_attendance_limit(student_id)
            
            if exceeded:
                # 이미 출석한 학생 (주 1회 초과)
                attendance_date = recent_dates[0] if recent_dates else ""
                flash(f'이번 주에 이미 출석했습니다. 출석일: {attendance_date}', 'danger')
                return redirect(url_for('attendance'))
        except Exception as e:
            # 확인 중 오류 발생 시 기존 방식으로 확인
            logging.error(f"출석 확인 중 오류: {e}")
            exceeded, count, recent_dates = check_weekly_attendance_limit(student_id)
            if exceeded:
                # 주간 1회 초과 출석 제한
                formatted_dates = ', '.join(recent_dates)
                flash(f'주간 출석 제한 (1회/주)을 초과했습니다. 최근 출석일: {formatted_dates}', 'danger')
                return redirect(url_for('attendance'))
        
        # name과 seat이 비어있는 경우 학생 정보 찾기
        if not name or not seat:
            student_data = load_student_data()
            student_info = student_data.get(student_id)
            
            # 학생 정보가 있으면 이름과 좌석 설정
            if student_info:
                name = student_info[0]
                seat = student_info[1]
            else:
                # 테스트용 하드코딩 데이터
                test_data = {
                    "10307": {"name": "박지호", "seat": "387"},
                    "20101": {"name": "강지훈", "seat": "331"},
                    "30107": {"name": "김리나", "seat": "175"},
                    "30207": {"name": "김유담", "seat": "281"},
                    "20240101": {"name": "홍길동", "seat": "A1"}
                }
                
                if student_id in test_data:
                    name = test_data[student_id]["name"]
                    seat = test_data[student_id]["seat"]
                else:
                    flash('해당 학번의 학생 정보를 찾을 수 없습니다.', 'danger')
                    return redirect(url_for('attendance'))
        
        # 마지막으로 한번 더 중복 출석 확인
        # 다른 탭이나 브라우저에서 동시에 요청이 들어올 경우 대비
        try:
            # 다시 한번 더 확인
            exceeded, count, recent_dates = check_weekly_attendance_limit(student_id)
            
            if exceeded:
                # 이미 출석한 학생
                attendance_date = recent_dates[0] if recent_dates else ""
                flash(f'이번 주에 이미 출석했습니다. 출석일: {attendance_date}', 'danger')
                return redirect(url_for('attendance'))
                
            # 출석 정보 저장
            # 교시 텍스트 설정
            period_text_for_db = period_text
            if period_text == "4교시 (도서실 이용 불가)":
                period_text_for_db = "4교시"
            
            # 출석 정보 저장
            if save_attendance(student_id, name, seat, period_text_for_db):
                # 캐시 갱신을 위해 캐시 키 삭제 (강제 새로고침)
                cache_key = f"weekly_limit_{student_id}" 
                if cache_key in attendance_status_cache:
                    del attendance_status_cache[cache_key]
                    
                flash('출석이 성공적으로 등록되었습니다!', 'success')
            else:
                flash('출석 등록에 실패했습니다.', 'danger')
        except Exception as e:
            flash(f'출석 등록 중 오류가 발생했습니다: {str(e)}', 'danger')
        
        return redirect(url_for('attendance'))
        
    # GET 요청 처리 (출석 폼 표시)
    return render_template('simple_form.html', 
                         now=now, 
                         period_text=period_text)

@app.route('/lookup_name')
def lookup_name():
    """학생 정보 조회 API"""
    student_id = request.args.get('student_id', '').strip()
    if not student_id:
        return jsonify({'error': '학번이 없습니다.'})
    
    # 학생 데이터 로드 (Excel 파일에서)
    student_data = load_student_data()
    student_info = student_data.get(student_id)
    
    if student_info:
        # Excel 파일에서 찾은 정보 반환
        return jsonify({
            'success': True,
            'name': student_info[0],
            'seat': student_info[1]
        })
    else:
        # 테스트용 하드코딩 데이터 (Excel 파일에서 찾지 못한 경우)
        test_data = {
            "10307": {"name": "박지호", "seat": "387"},
            "20101": {"name": "강지훈", "seat": "331"},
            "30107": {"name": "김리나", "seat": "175"},
            "30207": {"name": "김유담", "seat": "281"},
            "20240101": {"name": "홍길동", "seat": "A1"}
        }
        
        if student_id in test_data:
            return jsonify({
                'success': True,
                'name': test_data[student_id]["name"],
                'seat': test_data[student_id]["seat"]
            })
            
        # 어디에서도 찾지 못한 경우
        return jsonify({'error': '학번에 해당하는 학생 정보가 없습니다.'})

@app.route('/list', methods=['GET'])
def list_attendance():
    """List all attendance records"""
    if session.get('admin'):
        current_page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        sort_by = request.args.get('sort_by', 'date')  # 기본 정렬: 날짜
        sort_direction = request.args.get('sort_direction', 'desc')  # 기본 방향: 내림차순
        search_query = request.args.get('search', '').strip()
        search_field = request.args.get('search_field', 'all')
        
        # 모든 출석 기록 로드
        all_records = load_attendance()
        total_count = len(all_records)
        
        # 디버깅: 로드된 기록 수 로그
        logging.info(f"관리자 페이지 - 로드된 출석 기록: {total_count}개")
        
        # 검색 기능 적용
        if search_query:
            filtered_records = []
            search_query = search_query.lower()
            
            for record in all_records:
                # 검색 필드에 따른 필터링
                if search_field == 'all':
                    # 모든 필드에서 검색
                    if (
                        search_query in str(record.get('student_id', '')).lower() or
                        search_query in str(record.get('name', '')).lower() or
                        search_query in str(record.get('seat', '')).lower() or
                        search_query in str(record.get('period', '')).lower() or
                        search_query in str(record.get('date', '')).lower()
                    ):
                        filtered_records.append(record)
                else:
                    # 특정 필드에서만 검색
                    field_value = str(record.get(search_field, '')).lower()
                    if search_query in field_value:
                        filtered_records.append(record)
            
            all_records = filtered_records
            total_count = len(all_records)
        
        # 정렬 처리
        if sort_by and all_records:
            reverse = sort_direction.lower() == 'desc'
            
            def get_sort_key(record):
                value = record.get(sort_by)
                # 날짜는 특별 처리
                if sort_by == 'date':
                    try:
                        return datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                    except:
                        return datetime.min
                return str(value).lower() if value is not None else ""
                
            all_records = sorted(all_records, key=get_sort_key, reverse=reverse)
        
        # 간단한 페이지네이션
        total_pages = (len(all_records) + limit - 1) // limit if all_records else 1
        current_page = min(max(1, current_page), total_pages)
        start_idx = (current_page - 1) * limit
        end_idx = start_idx + limit
        paged_records = all_records[start_idx:end_idx] if all_records else []
        
        return render_template('list_simple.html', 
                              records=paged_records, 
                              current_page=current_page, 
                              total_pages=total_pages, 
                              total_count=total_count,
                              limit=limit,
                              sort_by=sort_by,
                              sort_direction=sort_direction,
                              search_query=search_query,
                              search_field=search_field)
    else:
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))

@app.route('/by_period')
def by_period():
    """교시별 출석 현황"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    # 날짜 선택 (기본값: 오늘)
    today = datetime.now(KST).strftime('%Y-%m-%d')
    selected_date = request.args.get('date', today)
    
    # 정렬 옵션 (기본: 좌석번호 오름차순)
    sort_by = request.args.get('sort_by', 'seat')
    sort_direction = request.args.get('sort_direction', 'asc')
    
    # 모든 기록 로드
    all_records = load_attendance()
    
    # 날짜로 필터링
    day_records = [r for r in all_records if r.get('date_only') == selected_date]
    
    # 교시별로 그룹화
    grouped_records = {}
    for record in day_records:
        period = record.get('period', '기타')
        if period not in grouped_records:
            grouped_records[period] = []
        grouped_records[period].append(record)
    
    # 교시 순서대로 정렬 (최근 교시가 상단에 오도록 역순 정렬)
    sorted_groups = {}
    period_order = ["10교시", "9교시", "8교시", "7교시", "6교시", "5교시", "4교시", "3교시", "2교시", "1교시", "시간 외", "기타"]
    
    for period in period_order:
        if period in grouped_records:
            # 각 교시 내에서 학생 정렬 적용
            students = grouped_records[period]
            
            # 정렬 기준에 따라 정렬
            if sort_by == 'seat':
                # 좌석번호 정렬 (숫자는 숫자끼리, 문자는 문자끼리)
                def seat_sort_key(student):
                    seat = student.get('seat', '')
                    if seat and seat.isdigit():
                        return (0, int(seat))  # 숫자는 앞에 배치
                    return (1, seat.lower())   # 문자는 뒤에 배치
                
                students = sorted(students, key=seat_sort_key, reverse=(sort_direction == 'desc'))
            elif sort_by == 'name':
                # 이름 정렬
                students = sorted(students, key=lambda s: s.get('name', ''), reverse=(sort_direction == 'desc'))
            elif sort_by == 'student_id':
                # 학번 정렬
                students = sorted(students, key=lambda s: s.get('student_id', ''), reverse=(sort_direction == 'desc'))
            
            sorted_groups[period] = students
    
    # 정의되지 않은 교시가 있다면 마지막에 추가
    for period in grouped_records:
        if period not in sorted_groups:
            sorted_groups[period] = grouped_records[period]
    
    # 날짜 포맷팅 (YYYY-MM-DD -> M월 D일)
    formatted_date = ''
    try:
        date_obj = datetime.strptime(selected_date, '%Y-%m-%d')
        weekday_names = ['월', '화', '수', '목', '금', '토', '일']
        weekday = weekday_names[date_obj.weekday()]
        formatted_date = f"{date_obj.month}월 {date_obj.day}일 ({weekday})"
    except:
        formatted_date = selected_date
    
    return render_template('by_period.html', 
                          grouped_records=sorted_groups, 
                          today=today,
                          selected_date=formatted_date,
                          sort_by=sort_by,
                          sort_direction=sort_direction)

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        access_id = request.form.get('access_id')
        password = request.form.get('password')
        
        admin_access_id = os.environ.get('ADMIN_ACCESS_ID', '20255008')
        admin_password = os.environ.get('ADMIN_PASSWORD', '1234')
        
        # 디버깅 로그
        logging.info(f"로그인 시도 - 입력 ID: {access_id}, 입력 PW: {password}")
        logging.info(f"설정된 ID: {admin_access_id}, 설정된 PW: {admin_password}")
        
        if access_id == admin_access_id and password == admin_password:
            session['admin'] = True
            flash('관리자로 로그인되었습니다.', 'success')
            return redirect(url_for('list_attendance'))  # 출석 목록으로 이동
        else:
            flash('접근 ID 또는 비밀번호가 올바르지 않습니다.', 'danger')
    
    if session.get('admin'):
        return redirect(url_for('list_attendance'))  # 출석 목록으로 이동
    
    return render_template('admin.html')

@app.route('/logout')
def logout():
    """Logout from admin"""
    session.pop('admin', None)
    flash('로그아웃되었습니다.', 'info')
    return redirect(url_for('attendance'))

@app.route('/admin_add_attendance', methods=['GET', 'POST'])
def admin_add_attendance():
    """관리자용 추가 출석 페이지"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    student_info = None
    attended = False
    last_attendance_date = None
    override = False
    
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        override_check = request.form.get('override_check') == 'on'
        
        # 학생 데이터 로드
        student_data = load_student_data()
        if student_id in student_data:
            name, seat = student_data[student_id]
            
            # 출석 체크를 위한 확인 로직
            # 이미 출석했는지 확인 (결과를 무시하고 attended 플래그만 추출)
            # 관리자 권한으로 추가 시 출석 제한을 무시하기 위해 admin_override=True 전달
            all_records = load_attendance()
            student_records = [r for r in all_records if r.get('student_id') == student_id]
            
            # 이번 주 출석 여부 확인
            sunday_str, saturday_str = get_current_week_range()
            sunday = datetime.strptime(sunday_str, '%Y-%m-%d')
            saturday = datetime.strptime(saturday_str, '%Y-%m-%d')
            
            # 이번주 출석 기록 필터링
            week_records = []
            for record in student_records:
                try:
                    record_date = datetime.strptime(record.get('date'), '%Y-%m-%d %H:%M:%S')
                    if sunday <= record_date <= saturday:
                        week_records.append(record)
                except:
                    pass
            
            attended = len(week_records) > 0
            if attended and week_records:
                # 가장 최근 출석일 찾기
                last_record = max(week_records, key=lambda r: datetime.strptime(r.get('date'), '%Y-%m-%d %H:%M:%S'))
                last_attendance_date = last_record.get('date').split(' ')[0]
                
            # 경고 여부 체크 (실제 구현은 여기에 추가)
            is_warned = False
            warning_info = None
            
            student_info = {
                'id': student_id,
                'name': name,
                'seat': seat,
                'is_warned': is_warned,
                'warning_info': warning_info
            }
            
            override = override_check
            
    return render_template('admin_add_attendance.html',
                          student_info=student_info,
                          attended=attended,
                          last_attendance_date=last_attendance_date,
                          override=override)

@app.route('/admin_add_attendance/confirm', methods=['POST'])
def admin_add_attendance_confirm():
    """관리자용 추가 출석 확인 처리"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    # 폼 데이터 가져오기
    student_id = request.form.get('student_id')
    name = request.form.get('name')
    seat = request.form.get('seat')
    override = request.form.get('override') == '1'
    
    if not student_id or not name or not seat:
        flash('필수 정보가 누락되었습니다.', 'danger')
        return redirect(url_for('admin_add_attendance'))
    
    try:
        # 관리자 권한으로 출석 추가
        current_period = get_current_period()
        period_text = "시간 외"
        
        if current_period == 0:
            period_text = "4교시"
        elif current_period > 0:
            period_text = f"{current_period}교시"
        
        # 중복 출석 체크 플래그 설정 (override = True면 중복 체크 무시)
        admin_override = override
        
        # 출석 기록 저장 (admin_override 파라미터 전달)
        save_attendance(student_id, name, seat, period_text, admin_override=admin_override)
        
        flash(f"✅ 관리자 권한으로 추가 출석이 완료되었습니다. 학번: {student_id}, 이름: {name}", "success")
        return redirect(url_for('by_period'))
        
    except Exception as e:
        flash(f'오류가 발생했습니다: {str(e)}', 'danger')
        return redirect(url_for('admin_add_attendance'))



@app.route('/admin/warnings')
def admin_warnings():
    """경고 학생 관리 페이지 (관리자만 접근 가능)"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    try:
        # 모든 경고 기록 가져오기
        warnings_ref = db.collection('warnings').order_by('warning_date', direction=firestore.Query.DESCENDING)
        warnings_docs = warnings_ref.get()
        
        warnings = []
        for doc in warnings_docs:
            warning_data = doc.to_dict()
            warning_data['id'] = doc.id
            
            # 날짜 변환
            if 'warning_date' in warning_data and warning_data['warning_date']:
                warning_data['warning_date'] = warning_data['warning_date'].replace(tzinfo=pytz.UTC).astimezone(KST)
            if 'expiry_date' in warning_data and warning_data['expiry_date']:
                warning_data['expiry_date'] = warning_data['expiry_date'].replace(tzinfo=pytz.UTC).astimezone(KST)
            
            # 경고가 만료되었는지 확인
            now = datetime.now(KST)
            warning_data['active'] = warning_data.get('active', False)
            if warning_data.get('expiry_date') and warning_data['expiry_date'] < now:
                warning_data['active'] = False
            
            warnings.append(warning_data)
        
        return render_template('admin_warnings.html', warnings=warnings)
    except Exception as e:
        flash(f'경고 정보를 불러오는 중 오류가 발생했습니다: {e}', 'danger')
        return redirect(url_for('list_attendance'))

@app.route('/admin/warnings/add', methods=['POST'])
def add_warning():
    """학생 경고 추가 처리"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    try:
        student_id = request.form.get('student_id')
        reason = request.form.get('reason', '도서실 이용 규정 위반')
        
        # 학생 정보 확인
        student_data = load_student_data().get(student_id)
        student_name = student_data[0] if student_data else None
        
        # 경고 기간 계산
        duration = request.form.get('duration')
        if duration == 'custom':
            days = int(request.form.get('customDuration', 7))
        else:
            days = int(duration)
        
        warning_date = datetime.now(KST)
        expiry_date = warning_date + timedelta(days=days)
        
        # Firestore에 경고 추가
        warning_data = {
            'student_id': student_id,
            'name': student_name,
            'reason': reason,
            'warning_date': warning_date,
            'expiry_date': expiry_date,
            'active': True,
            'created_at': firestore.SERVER_TIMESTAMP
        }
        
        db.collection('warnings').add(warning_data)
        
        flash(f'학생(학번: {student_id})에게 {days}일 동안의 경고가 추가되었습니다.', 'success')
    except Exception as e:
        flash(f'경고 추가 중 오류가 발생했습니다: {e}', 'danger')
    
    return redirect(url_for('admin_warnings'))

@app.route('/admin/warnings/remove/<warning_id>')
def remove_warning(warning_id):
    """학생 경고 해제 처리"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    try:
        # 경고 정보 가져오기
        warning_ref = db.collection('warnings').document(warning_id)
        warning_doc = warning_ref.get()
        
        if not warning_doc.exists:
            flash('해당 경고를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('admin_warnings'))
        
        warning_data = warning_doc.to_dict()
        student_id = warning_data.get('student_id')
        
        # 경고 비활성화 (삭제하지 않고 비활성화만 함)
        warning_ref.update({
            'active': False,
            'updated_at': firestore.SERVER_TIMESTAMP
        })
        
        flash(f'학생(학번: {student_id})의 경고가 해제되었습니다.', 'success')
    except Exception as e:
        flash(f'경고 해제 중 오류가 발생했습니다: {e}', 'danger')
    
    return redirect(url_for('admin_warnings'))

@app.route('/admin/warnings/delete/<warning_id>')
def delete_warning(warning_id):
    """학생 경고 완전 삭제 처리"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    try:
        # 경고 정보 가져오기
        warning_ref = db.collection('warnings').document(warning_id)
        warning_doc = warning_ref.get()
        
        if not warning_doc.exists:
            flash('해당 경고를 찾을 수 없습니다.', 'danger')
            return redirect(url_for('admin_warnings'))
        
        warning_data = warning_doc.to_dict()
        student_id = warning_data.get('student_id')
        
        # 경고 완전 삭제
        warning_ref.delete()
        
        flash(f'학생(학번: {student_id})의 경고가 완전히 삭제되었습니다.', 'success')
    except Exception as e:
        flash(f'경고 삭제 중 오류가 발생했습니다: {e}', 'danger')
    
    return redirect(url_for('admin_warnings'))

@app.route('/admin/warnings/delete-all')
def delete_all_warnings():
    """모든 경고 삭제 처리"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    try:
        # 모든 경고 가져오기
        warnings_ref = db.collection('warnings').get()
        
        # 배치 삭제 (한 번에 최대 500개)
        batch_size = 0
        batch = db.batch()
        
        for doc in warnings_ref:
            batch.delete(doc.reference)
            batch_size += 1
            
            # 배치 크기가 500에 도달하면 커밋
            if batch_size >= 500:
                batch.commit()
                batch = db.batch()
                batch_size = 0
        
        # 남은 배치 커밋
        if batch_size > 0:
            batch.commit()
        
        flash('모든 경고가 성공적으로 삭제되었습니다.', 'success')
    except Exception as e:
        flash(f'경고 삭제 중 오류가 발생했습니다: {e}', 'danger')
    
    return redirect(url_for('admin_warnings'))

@app.route('/edit_seat')
def edit_seat():
    """학생 좌석번호 수정 페이지 (관리자 전용)"""
    if not session.get('admin'):
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('admin_login'))
    return send_from_directory('static', 'edit_seat.html')

@app.route('/bulk_edit_seats')
def bulk_edit_seats():
    """학생 좌석번호 일괄 수정 페이지 (관리자 전용)"""
    if not session.get('admin'):
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('admin_login'))
    return render_template('bulk_edit_seats.html')

@app.route('/update_seat', methods=['POST'])
def update_seat():
    """학생 좌석번호 업데이트 API (관리자 전용)"""
    if not session.get('admin'):
        return jsonify({"error": "관리자 권한이 필요합니다."}), 403
    
    student_id = request.json.get('student_id')
    new_seat = request.json.get('new_seat')
    
    if not student_id or not new_seat:
        return jsonify({"error": "학번과 새 좌석번호를 모두 입력해주세요."}), 400
    
    try:
        # Excel 파일 수정
        import pandas as pd
        from openpyxl import load_workbook
        
        excel_path = 'students.xlsx'
        
        # pandas 미리보기로 열 이름 확인
        df_preview = pd.read_excel(excel_path, nrows=1)
        column_names = df_preview.columns.tolist()
        
        # 학번과 좌석번호 열 확인
        id_column = [col for col in column_names if '학번' in col or 'ID' in col.upper()][0]
        seat_column = [col for col in column_names if '좌석' in col or 'SEAT' in col.upper()][0]
        
        # openpyxl로 직접 셀 수정
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 학번 열 및 좌석번호 열의 인덱스 찾기
        col_indices = {}
        for idx, col in enumerate(ws[1]):
            if col.value == id_column:
                col_indices['id'] = idx + 1  # 1-based index
            elif col.value == seat_column:
                col_indices['seat'] = idx + 1  # 1-based index
        
        # 학번으로 학생 찾아 좌석번호 수정
        student_found = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # 2행부터 시작 (헤더 제외)
            cell_id = row[col_indices['id'] - 1].value  # 0-based index
            if str(cell_id) == student_id:
                student_found = True
                old_seat = row[col_indices['seat'] - 1].value
                # 새 좌석번호로 업데이트
                ws.cell(row=row_idx, column=col_indices['seat']).value = new_seat
                break
        
        if not student_found:
            return jsonify({"error": f"학번 {student_id}를 찾을 수 없습니다."}), 404
        
        # 변경 사항 저장
        wb.save(excel_path)
        
        # 학생 데이터 캐시 초기화 (새로고침)
        global _student_data_cache, _student_data_timestamp
        _student_data_cache = None
        _student_data_timestamp = None
        
        return jsonify({
            "success": True, 
            "message": f"학번 {student_id}의 좌석번호가 {old_seat}에서 {new_seat}로 업데이트되었습니다."
        }), 200
    except Exception as e:
        return jsonify({"error": f"좌석번호 업데이트 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/add_student_form', methods=['GET'])
def add_student_form():
    """학생 추가 폼 페이지"""
    if not session.get('admin'):
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('admin_login'))
    
    return render_template('add_student.html')
    
@app.route('/delete_student_form', methods=['GET'])
def delete_student_form():
    """학생 삭제 폼 페이지"""
    if not session.get('admin'):
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('admin_login'))
    
    return render_template('delete_student.html')

@app.route('/api/add_direct_student', methods=['POST'])
def add_direct_student():
    """새 학생을 직접 추가하는 API"""
    if not session.get('admin'):
        return jsonify({"error": "관리자 권한이 필요합니다."}), 403
    
    data = request.json
    if not data or 'student_id' not in data or 'name' not in data or 'seat' not in data:
        return jsonify({"error": "필수 정보가 누락되었습니다. (학번, 이름, 좌석번호)"}), 400
    
    student_id = data['student_id']
    name = data['name']
    seat = data['seat']
    
    try:
        # 학생 추가 모듈 사용
        from add_student import add_new_student
        success, message = add_new_student(student_id, name, seat)
        
        if success:
            # 데이터를 강제로 새로 로딩하여 캐시 갱신
            global _student_data_cache
            _student_data_cache = None
            load_student_data(force_reload=True)
            
            return jsonify({
                "success": True,
                "message": message
            })
        else:
            return jsonify({
                "warning": message
            }), 200
            
    except Exception as e:
        return jsonify({"error": f"학생 추가 중 오류 발생: {str(e)}"}), 500
        
@app.route('/api/delete_student', methods=['POST'])
def delete_student_api():
    """학생 삭제 API"""
    if not session.get('admin'):
        return jsonify({"error": "관리자 권한이 필요합니다."}), 403
    
    data = request.json
    if not data or 'student_id' not in data:
        return jsonify({"error": "학번이 필요합니다."}), 400
    
    student_id = data['student_id']
    
    try:
        # 학생 삭제 모듈 사용
        from delete_student import delete_student
        success, message = delete_student(student_id)
        
        if success:
            # 데이터를 강제로 새로 로딩하여 캐시 갱신
            global _student_data_cache
            _student_data_cache = None
            load_student_data(force_reload=True)
            
            return jsonify({
                "success": True,
                "message": message
            })
        else:
            return jsonify({
                "warning": message
            }), 200
            
    except Exception as e:
        return jsonify({"error": f"학생 삭제 중 오류 발생: {str(e)}"}), 500

@app.route('/api/students', methods=['GET'])
def api_get_students():
    """학생 정보 조회 API (학번 목록으로 학생 정보 반환)"""
    if not session.get('admin'):
        return jsonify({"error": "관리자 권한이 필요합니다."}), 403
    
    # 쿼리 파라미터에서 학번 목록 가져오기
    student_ids = request.args.get('ids', '')
    if not student_ids:
        return jsonify([])
    
    student_id_list = student_ids.split(',')
    
    try:
        # 학생 데이터 로드
        students_data = load_student_data()
        
        # 새 학생 텍스트 파일이 있으면 추가 로드
        new_students = {}
        if os.path.exists('new_students.txt'):
            with open('new_students.txt', 'r', encoding='utf-8') as f:
                for line in f:
                    parts = line.strip().split(',')
                    if len(parts) >= 3:
                        new_students[parts[0]] = (parts[1], parts[2])
        
        # 새로운 학생 데이터도 추가하기 위해 현재 요청 목록에서 학번을 모두 처리
        result = []
        for student_id in student_id_list:
            if student_id in students_data:
                name, seat = students_data[student_id]
                result.append({
                    "student_id": student_id,
                    "name": name or "이름 없음",  # 이름이 없는 경우 대체 텍스트
                    "seat": seat or "-"           # 좌석번호가 없는 경우 대체 텍스트
                })
            elif student_id in new_students:
                name, seat = new_students[student_id]
                result.append({
                    "student_id": student_id,
                    "name": name or "이름 없음",
                    "seat": seat or "-"
                })
            else:
                # 데이터베이스에 없는 새 학생인 경우
                result.append({
                    "student_id": student_id,
                    "name": "새 학생",
                    "seat": "-"
                })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": f"학생 정보 조회 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/api/bulk_update_seats', methods=['POST'])
def api_bulk_update_seats():
    """학생 좌석번호 및 이름 일괄 업데이트 API (관리자 전용)"""
    if not session.get('admin'):
        return jsonify({"error": "관리자 권한이 필요합니다."}), 403
    
    data = request.json
    if not data or 'changes' not in data or not data['changes']:
        return jsonify({"error": "유효한 데이터가 없습니다."}), 400
    
    try:
        # Excel 파일 업데이트
        import pandas as pd
        from openpyxl import load_workbook
        
        excel_path = 'students.xlsx'
        
        # pandas로 열 이름 확인
        df_preview = pd.read_excel(excel_path, nrows=1)
        column_names = df_preview.columns.tolist()
        
        # 학번과 좌석번호 열 확인
        id_column = [col for col in column_names if '학번' in col or 'ID' in col.upper()][0]
        seat_column = [col for col in column_names if '좌석' in col or 'SEAT' in col.upper()][0]
        
        # openpyxl로 직접 셀 수정
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 학번 열 및 좌석번호 열의 인덱스 찾기
        col_indices = {}
        for idx, col in enumerate(ws[1]):
            if col.value == id_column:
                col_indices['id'] = idx + 1  # 1-based index
            elif col.value == seat_column:
                col_indices['seat'] = idx + 1  # 1-based index
        
        # 변경 사항 추적
        changes_count = 0
        not_found_count = 0
        
        # 학번별 새 좌석번호와 이름 매핑 생성
        changes_map = {}
        for item in data['changes']:
            student_id = item['student_id']
            changes_map[student_id] = {
                'new_seat': item['new_seat']
            }
            # 새 이름이 있는 경우만 추가
            if 'new_name' in item and item['new_name']:
                changes_map[student_id]['new_name'] = item['new_name']
        
        # 엑셀 파일 수정
        seat_changes_count = 0
        name_changes_count = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # 2행부터 시작 (헤더 제외)
            cell_id = row[col_indices['id'] - 1].value  # 0-based index
            if str(cell_id) in changes_map:
                change_item = changes_map[str(cell_id)]
                
                # 좌석번호 업데이트
                ws.cell(row=row_idx, column=col_indices['seat']).value = change_item['new_seat']
                seat_changes_count += 1
                
                # 이름 업데이트 (이름 칼럼이 존재하고 새 이름이 제공된 경우)
                if 'name' in col_indices and 'new_name' in change_item:
                    ws.cell(row=row_idx, column=col_indices['name']).value = change_item['new_name']
                    name_changes_count += 1
                
                changes_count = seat_changes_count + name_changes_count
                
                # 처리한 항목 제거
                del changes_map[str(cell_id)]
        
        # 데이터베이스에 없는 새 학생 추가
        new_students_count = 0
        
        if changes_map and 'name' in col_indices:  # 이름 열이 있는 경우만 새 학생 추가 가능
            # 마지막 행 다음에 새 학생 추가
            last_row = ws.max_row + 1
            
            for student_id, change_item in list(changes_map.items()):
                if 'new_name' in change_item:  # 이름이 제공된 경우에만 새 학생으로 추가
                    # 새 학생 정보 추가
                    ws.cell(row=last_row, column=col_indices['id']).value = student_id
                    ws.cell(row=last_row, column=col_indices['seat']).value = change_item['new_seat']
                    ws.cell(row=last_row, column=col_indices['name']).value = change_item['new_name']
                    
                    new_students_count += 1
                    last_row += 1
                    
                    # 처리한 항목 제거
                    del changes_map[student_id]
        
        # 미처리된 학번 수 계산
        not_found_count = len(changes_map)
        
        # 변경 사항 저장
        wb.save(excel_path)
        
        # 학생 데이터 캐시 초기화
        global _student_data_cache, _student_data_timestamp
        _student_data_cache = None
        _student_data_timestamp = None
        
        return jsonify({
            "success": True,
            "message": f"성공적으로 업데이트되었습니다: 좌석번호 {seat_changes_count}개, 이름 {name_changes_count}개 변경됨, 새 학생 {new_students_count}명 추가됨. {not_found_count}개의 학번은 처리할 수 없습니다."
        })
    except Exception as e:
        return jsonify({"error": f"학생 정보 일괄 업데이트 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/delete_records', methods=['POST'])
def delete_records():
    """Delete selected attendance records (admin only)"""
    if not session.get('admin'):
        flash('관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('admin_login'))
    
    # 이전 페이지 확인 (교시별 보기 또는 출석 목록)
    referrer = request.referrer
    redirect_to_period_view = False
    
    # referrer URL에서 경로 추출 (by_period인지 확인)
    if referrer:
        from urllib.parse import urlparse
        parsed_uri = urlparse(referrer)
        path = parsed_uri.path
        if path.endswith('/by_period'):
            redirect_to_period_view = True
            # 쿼리 파라미터도 기억 (날짜 등)
            query = parsed_uri.query
    
    record_ids = request.form.getlist('record_ids[]')
    if not record_ids:
        flash('삭제할 기록을 선택해주세요.', 'warning')
        if redirect_to_period_view:
            return redirect(url_for('by_period'))
        return redirect(url_for('list_attendance'))
    
    try:
        # CSV 파일에서 삭제
        df = pd.read_csv('attendance.csv', encoding='utf-8')
        
        # record_ids는 실제로는 인덱스 번호들
        indices_to_delete = []
        for record_id in record_ids:
            try:
                idx = int(record_id)
                if 0 <= idx < len(df):
                    indices_to_delete.append(idx)
            except ValueError:
                continue
        
        if indices_to_delete:
            # 인덱스 기준으로 행 삭제
            df = df.drop(indices_to_delete)
            # CSV 파일 다시 저장
            df.to_csv('attendance.csv', index=False, encoding='utf-8')
            
            # Firebase에서도 삭제 (해당 기록이 있는 경우)
            if db:
                for idx in indices_to_delete:
                    try:
                        # 원본 데이터에서 학번과 날짜 정보 추출하여 Firebase에서도 삭제
                        original_df = pd.read_csv('attendance.csv', encoding='utf-8')
                        if idx < len(original_df):
                            student_id = str(original_df.iloc[idx]['학번'])
                            date_str = original_df.iloc[idx]['날짜']
                            
                            # Firebase에서 삭제
                            db.collection('attendance').document(student_id).collection('records').document(date_str).delete()
                    except Exception as firebase_error:
                        logging.error(f"Firebase 삭제 중 오류: {firebase_error}")
        
        flash(f'{len(indices_to_delete)}개의 기록이 삭제되었습니다.', 'success')
    except Exception as e:
        flash(f'기록 삭제 중 오류가 발생했습니다: {e}', 'danger')
    
    # 이전 페이지가 교시별 보기였으면 그쪽으로 리다이렉트
    if redirect_to_period_view:
        # 원래 URL에 있던 쿼리 파라미터(예: 날짜) 유지
        if 'query' in locals() and query:
            return redirect(url_for('by_period') + '?' + query)
        return redirect(url_for('by_period'))
    
    return redirect(url_for('list_attendance'))

@app.route('/stats')
def stats():
    """학생 통계 페이지 - 요일별, 교시별, 학생별 출석 통계"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    # 기본 날짜 설정 (기본: 지난 30일)
    today = datetime.now(KST).date()
    default_start_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
    default_end_date = today.strftime('%Y-%m-%d')
    
    # URL 파라미터에서 날짜와 보기 모드 가져오기
    start_date = request.args.get('start_date', default_start_date)
    end_date = request.args.get('end_date', default_end_date)
    view_mode = request.args.get('view_mode', 'total')  # 'total' 또는 'weekly'
    
    # 날짜 문자열을 datetime 객체로 변환
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # 출석 기록 불러오기
    records = load_attendance()
    
    # 기간 내 출석 기록만 필터링
    filtered_records = []
    for record in records:
        try:
            # 문자열 날짜를 datetime 객체로 변환
            record_date_str = record.get('date', '').split()[0]  # '2025-05-22' 형식으로 추출
            record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date()
            
            # 날짜 범위 확인
            if start_date_obj <= record_date <= end_date_obj:
                filtered_records.append(record)
        except (ValueError, AttributeError, IndexError):
            continue
            
    # 주차별 데이터 구성 (view_mode가 'weekly'일 때 사용)
    weekly_stats = []
    
    if view_mode == 'weekly':
        # 시작일부터 종료일까지의 주차별 데이터 구성
        current_week_start = start_date_obj - timedelta(days=start_date_obj.weekday())  # 해당 주의 월요일
        
        while current_week_start <= end_date_obj:
            current_week_end = current_week_start + timedelta(days=6)  # 주의 일요일
            
            # 이번 주 레이블 (월/일 ~ 월/일)
            week_label = f"{current_week_start.month}/{current_week_start.day} ~ {current_week_end.month}/{current_week_end.day}"
            
            # 이번 주 기록 필터링
            weekly_records = []
            for record in filtered_records:
                try:
                    record_date_str = record.get('date', '').split()[0]
                    record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date()
                    if current_week_start <= record_date <= current_week_end:
                        weekly_records.append(record)
                except (ValueError, AttributeError):
                    continue
            
            # 이번 주 요일별 교시 데이터 준비
            week_data = {
                'label': week_label,
                'start_date': current_week_start,
                'end_date': current_week_end,
                'records_count': len(weekly_records),
                'weekday_period_data': {}
            }
            
            # 요일별 교시 카운팅
            all_periods_week = set()
            weekday_period_data_week = {}
            
            for record in weekly_records:
                try:
                    record_date_str = record.get('date', '').split()[0]
                    record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date()
                    weekday = record_date.weekday()
                    weekday_name = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일'][weekday]
                    period = record.get('period', '미지정')
                    
                    # 모든 교시 목록 수집
                    all_periods_week.add(period)
                    
                    # 요일별 교시 데이터 추가
                    if weekday_name not in weekday_period_data_week:
                        weekday_period_data_week[weekday_name] = Counter()
                    weekday_period_data_week[weekday_name][period] += 1
                    
                except (ValueError, AttributeError, IndexError):
                    continue
            
            # 교시 정렬 함수
            def period_sort_key_week(period):
                if '교시' in period:
                    try:
                        return int(period.split('교시')[0])
                    except ValueError:
                        return 999
                elif period == '시간 외':
                    return 998
                else:
                    return 999
            
            # 정렬된 교시 목록
            sorted_periods_week = sorted(all_periods_week, key=period_sort_key_week)
            
            # 요일 순서대로 정렬
            weekday_order = {'월요일': 0, '화요일': 1, '수요일': 2, '목요일': 3, '금요일': 4, '토요일': 5, '일요일': 6}
            
            # 요일별 교시 데이터 정렬
            weekday_period_stats_week = []
            for weekday_name in sorted(weekday_period_data_week.keys(), key=lambda x: weekday_order.get(x, 7)):
                period_counts = []
                for period in sorted_periods_week:
                    count = weekday_period_data_week.get(weekday_name, Counter()).get(period, 0)
                    period_counts.append((period, count))
                weekday_period_stats_week.append((weekday_name, period_counts))
            
            week_data['weekday_period_stats'] = weekday_period_stats_week
            week_data['sorted_periods'] = sorted_periods_week
            
            weekly_stats.append(week_data)
            
            # 다음 주로 이동
            current_week_start = current_week_end + timedelta(days=1)
    
    # 총 방문자 수
    total_visitors = len(filtered_records)
    
    # 요일별 통계
    weekday_counter = Counter()
    
    # 요일별 교시 통계를 위한 딕셔너리 (2차원 데이터)
    weekday_period_data = {}
    all_periods = set()
    
    for record in filtered_records:
        try:
            record_date_str = record.get('date', '').split()[0]
            record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date()
            # 요일 이름 (0:월요일 ~ 6:일요일로 변환)
            weekday = record_date.weekday()
            weekday_name = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일'][weekday]
            period = record.get('period', '미지정')
            
            # 요일별 카운터 증가
            weekday_counter[weekday_name] += 1
            
            # 모든 교시 목록 수집
            all_periods.add(period)
            
            # 요일별 교시 데이터 추가
            if weekday_name not in weekday_period_data:
                weekday_period_data[weekday_name] = Counter()
            weekday_period_data[weekday_name][period] += 1
            
        except (ValueError, AttributeError, IndexError):
            continue
    
    # 요일 순서대로 정렬
    weekday_order = {'월요일': 0, '화요일': 1, '수요일': 2, '목요일': 3, '금요일': 4, '토요일': 5, '일요일': 6}
    weekday_stats = sorted(weekday_counter.items(), key=lambda x: weekday_order.get(x[0], 7))
    
    # 최대 요일 카운트 (그래프 비율용)
    max_day_count = max(weekday_counter.values()) if weekday_counter else 1
    
    # 교시 정렬 함수
    def period_sort_key(period):
        if '교시' in period:
            try:
                return int(period.split('교시')[0])
            except ValueError:
                return 999
        elif period == '시간 외':
            return 998
        else:
            return 999
    
    # 정렬된 교시 목록 생성
    sorted_periods = sorted(all_periods, key=period_sort_key)
    
    # 요일별 교시 데이터 정렬
    weekday_period_stats = []
    for weekday_name, _ in weekday_stats:  # 요일 순서 유지
        period_counts = []
        for period in sorted_periods:
            count = weekday_period_data.get(weekday_name, Counter()).get(period, 0)
            period_counts.append((period, count))
        weekday_period_stats.append((weekday_name, period_counts))
    
    # 교시별 통계
    period_counter = Counter()
    for record in filtered_records:
        period = record.get('period', '미지정')
        period_counter[period] += 1
    
    # 교시 순서대로 정렬 (1교시부터 오름차순)
    def period_sort_key(item):
        period = item[0]
        if '교시' in period:
            try:
                return int(period.split('교시')[0])
            except ValueError:
                return 999
        elif period == '시간 외':
            return 998
        else:
            return 999
    
    period_stats = sorted(period_counter.items(), key=period_sort_key)
    
    # 최대 교시 카운트 (그래프 비율용)
    max_period_count = max(period_counter.values()) if period_counter else 1
    
    # 학생별 통계 (방문 빈도 상위 10명)
    student_counter = Counter()
    student_names = {}  # 학번별 이름 저장
    
    for record in filtered_records:
        student_id = record.get('student_id', '')
        name = record.get('name', '')
        if student_id:
            student_counter[student_id] += 1
            student_names[student_id] = name
    
    # 상위 10명 학생 추출
    top_students = []
    for student_id, count in student_counter.most_common(10):
        top_students.append({
            'student_id': student_id,
            'name': student_names.get(student_id, ''),
            'count': count
        })
    
    # 최대 학생 방문 횟수 (그래프 비율용)
    max_student_count = max(student_counter.values()) if student_counter else 1
    
    return render_template('stats.html', 
                           start_date=start_date,
                           end_date=end_date,
                           total_visitors=total_visitors,
                           weekday_stats=weekday_stats,
                           max_day_count=max_day_count,
                           period_stats=period_stats,
                           max_period_count=max_period_count,
                           top_students=top_students,
                           max_student_count=max_student_count,
                           weekday_period_stats=weekday_period_stats,
                           sorted_periods=sorted_periods,
                           view_mode=view_mode,
                           weekly_stats=weekly_stats)

@app.route('/health')
def health():
    return 'OK', 200

@app.route('/debug_firebase')
def debug_firebase():
    """Firebase 디버그 페이지 (관리자만)"""
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    if not db:
        return "<h1>Firebase 연결이 없습니다</h1>"
    
    results = []
    kst = pytz.timezone('Asia/Seoul')
    now = datetime.now(kst)
    date_str = now.strftime('%Y-%m-%d')
    
    # 1. Test write operation
    try:
        test_data = {
            'student_id': 'DEBUG001',
            'name': '디버그테스트',
            'seat': 'D01',
            'period': '디버그교시',
            'date': now.strftime('%Y-%m-%d %H:%M:%S'),
            'date_only': date_str,
            'timestamp': firestore.SERVER_TIMESTAMP
        }
        
        # Write to admin collection
        admin_ref = db.collection('admin').document(f"{date_str}_디버그교시").collection('students').document('DEBUG001')
        admin_ref.set(test_data)
        results.append("✅ admin 컬렉션 저장 성공")
        
        # Write to attendance collection
        attendance_ref = db.collection('attendance').document('DEBUG001').collection('records').document(date_str)
        attendance_ref.set(test_data)
        results.append("✅ attendance 컬렉션 저장 성공")
        
    except Exception as e:
        results.append(f"❌ 저장 실패: {e}")
    
    # 2. Test read operations
    try:
        # Read from admin
        admin_docs = list(db.collection('admin').stream())
        results.append(f"📂 admin 컬렉션: {len(admin_docs)}개 문서")
        
        for doc in admin_docs[:3]:
            students = list(doc.reference.collection('students').stream())
            results.append(f"  - {doc.id}: {len(students)}명")
            for student in students[:2]:
                data = student.to_dict()
                results.append(f"    └─ {student.id}: {data.get('name')}")
            
        # Read from attendance
        attendance_docs = list(db.collection('attendance').stream())
        results.append(f"📂 attendance 컬렉션: {len(attendance_docs)}개 학생")
        
        for doc in attendance_docs[:3]:
            records = list(doc.reference.collection('records').stream())
            results.append(f"  - 학생 {doc.id}: {len(records)}개 기록")
            for record in records[:2]:
                data = record.to_dict()
                results.append(f"    └─ {record.id}: {data.get('name')} - {data.get('period')}")
            
    except Exception as e:
        results.append(f"❌ 읽기 실패: {e}")
    
    # 3. All collections
    try:
        all_collections = list(db.collections())
        results.append(f"📚 전체 컬렉션: {[c.id for c in all_collections]}")
    except Exception as e:
        results.append(f"❌ 컬렉션 목록 실패: {e}")
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Firebase 디버그</title>
        <style>
            body {{ font-family: monospace; background: #1a1a1a; color: #fff; padding: 20px; }}
            pre {{ background: #2d2d2d; padding: 15px; border-radius: 5px; }}
        </style>
    </head>
    <body>
        <h1>Firebase 디버그 결과</h1>
        <pre>{'<br>'.join(results)}</pre>
        <p><a href="/list">← 출석 목록으로 돌아가기</a></p>
    </body>
    </html>
    """
    
    return html

@app.route('/add_sample_data')
def add_sample_data():
    """샘플 출석 데이터 추가 (관리자만)"""
    if not session.get('admin'):
        flash('관리자 로그인이 필요합니다.', 'warning')
        return redirect(url_for('admin_login'))
    
    try:
        if not db:
            flash('Firebase 연결 실패', 'danger')
            return redirect(url_for('list_attendance'))
        
        # 한국 시간대 설정
        kst = pytz.timezone('Asia/Seoul')
        now = datetime.now(kst)
        
        # 오늘 날짜 문자열
        date_str = now.strftime('%Y-%m-%d')
        datetime_str = now.strftime('%Y-%m-%d %H:%M:%S')
        
        # 샘플 학생 데이터
        sample_students = [
            {"student_id": "10307", "name": "박지호", "seat": "387", "period": "1교시"},
            {"student_id": "20101", "name": "강지훈", "seat": "331", "period": "2교시"},
            {"student_id": "30107", "name": "김리나", "seat": "175", "period": "3교시"},
            {"student_id": "30207", "name": "김유담", "seat": "281", "period": "1교시"},
            {"student_id": "20240101", "name": "홍길동", "seat": "A1", "period": "시간 외"}
        ]
        
        count = 0
        for student in sample_students:
            try:
                # 출석 데이터 구조
                attendance_data = {
                    'student_id': student['student_id'],
                    'name': student['name'],
                    'seat': student['seat'],
                    'period': student['period'],
                    'date': datetime_str,
                    'date_only': date_str,
                    'timestamp': firestore.SERVER_TIMESTAMP
                }
                
                # 1. attendance/{student_id}/records/{date} 구조에 저장
                student_ref = db.collection('attendance').document(student['student_id']).collection('records').document(date_str)
                student_ref.set(attendance_data)
                
                # 2. admin/{date_period}/students/{student_id} 구조에 저장
                date_period_key = f"{date_str}_{student['period']}"
                admin_ref = db.collection('admin').document(date_period_key).collection('students').document(student['student_id'])
                admin_ref.set(attendance_data)
                
                count += 1
                logging.info(f"샘플 데이터 추가: {student['name']} ({student['student_id']}) - {student['period']}")
                
            except Exception as e:
                logging.error(f"데이터 추가 실패 - {student['name']}: {e}")
        
        flash(f'샘플 데이터 {count}개를 성공적으로 추가했습니다.', 'success')
        
    except Exception as e:
        flash(f'샘플 데이터 추가 실패: {str(e)}', 'danger')
        logging.error(f"샘플 데이터 추가 중 오류: {e}")
    
    return redirect(url_for('list_attendance'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
