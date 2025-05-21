"""
새 학생 정보를 삭제하는 간단한 스크립트
- 엑셀 파일에서 학생 정보를 찾아 삭제
"""
import openpyxl
import os

def delete_student(student_id):
    """학생 정보를 Excel 파일에서 삭제"""
    excel_file = 'students.xlsx'
    
    # 엑셀 파일이 없는 경우 에러 반환
    if not os.path.exists(excel_file):
        return False, "학생 데이터 파일이 없습니다."
    
    try:
        # 엑셀 파일 읽기
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 삭제할 행 찾기
        row_to_delete = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and str(row[0]) == str(student_id):
                row_to_delete = row_idx
                break
        
        # 학생 정보를 찾지 못한 경우
        if not row_to_delete:
            return False, f"학번 {student_id}의 학생 정보를 찾을 수 없습니다."
        
        # 행 삭제 대신 학생 정보 비우기 (행 삭제시 인덱스 문제 발생)
        student_name = ws.cell(row=row_to_delete, column=2).value
        ws.cell(row=row_to_delete, column=1).value = None  # 학번 비우기
        ws.cell(row=row_to_delete, column=2).value = None  # 이름 비우기
        ws.cell(row=row_to_delete, column=3).value = None  # 좌석번호 비우기
        
        # 파일 저장
        wb.save(excel_file)
        return True, f"학생 삭제 성공: {student_id}, {student_name}"
    except Exception as e:
        return False, f"삭제 중 오류: {str(e)}"

if __name__ == "__main__":
    # 테스트 코드
    success, message = delete_student("10701")
    print(message)