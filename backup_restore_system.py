#!/usr/bin/env python3
"""
Replit 재부팅 시 데이터 영구 보존을 위한 백업/복원 시스템
Firebase를 사용하여 학생 데이터와 출석 데이터를 자동 백업/복원
"""

import os
import json
import logging
from datetime import datetime
import pytz
import openpyxl
from openpyxl import Workbook

# Firebase 관련 라이브러리
import firebase_admin
from firebase_admin import credentials, firestore

# 한국 시간대 설정
KST = pytz.timezone('Asia/Seoul')

def init_firebase():
    """Firebase 초기화"""
    try:
        FIREBASE_CREDENTIALS_JSON = os.environ.get("FIREBASE_CREDENTIALS_JSON")
        if not FIREBASE_CREDENTIALS_JSON:
            raise ValueError("FIREBASE_CREDENTIALS_JSON 환경변수가 없습니다.")
        
        # JSON 문자열을 딕셔너리로 변환
        firebase_config = json.loads(FIREBASE_CREDENTIALS_JSON)
        
        # Firebase 인증 정보 생성
        cred = credentials.Certificate(firebase_config)
        
        # Firebase 앱 초기화 (이미 초기화되었는지 확인)
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
        
        # Firestore 클라이언트 설정
        db = firestore.client()
        print("✅ Firebase 초기화 성공")
        return db
    except Exception as e:
        print(f"❌ Firebase 초기화 오류: {e}")
        return None

def backup_students_to_firebase(db, excel_file='students.xlsx'):
    """
    students.xlsx 파일을 Firebase에 백업
    """
    try:
        if not os.path.exists(excel_file):
            print(f"❌ {excel_file} 파일이 없습니다.")
            return False
            
        # Excel 파일 읽기
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 헤더 읽기
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"📋 컬럼: {headers}")
        
        # 학생 데이터 수집
        students_data = []
        student_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # 첫 번째 컬럼(이름 또는 학번)이 있는 경우만
                student_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        student_dict[header] = str(row[i]) if row[i] is not None else ''
                students_data.append(student_dict)
                student_count += 1
        
        # Firebase에 백업
        backup_doc = {
            'backup_date': datetime.now(KST),
            'student_count': student_count,
            'headers': headers,
            'students': students_data,
            'file_name': excel_file
        }
        
        # 백업 저장
        backup_ref = db.collection('backups').document('students_backup')
        backup_ref.set(backup_doc)
        
        print(f"✅ 학생 데이터 {student_count}명 Firebase 백업 완료")
        return True
        
    except Exception as e:
        print(f"❌ 학생 데이터 백업 실패: {e}")
        return False

def restore_students_from_firebase(db, excel_file='students.xlsx'):
    """
    Firebase에서 학생 데이터를 복원하여 Excel 파일 생성
    """
    try:
        # 백업 데이터 가져오기
        backup_ref = db.collection('backups').document('students_backup')
        backup_doc = backup_ref.get()
        
        if not backup_doc.exists:
            print("❌ Firebase에 학생 데이터 백업이 없습니다.")
            return False
        
        backup_data = backup_doc.to_dict()
        students = backup_data.get('students', [])
        headers = backup_data.get('headers', [])
        
        if not students:
            print("❌ 백업된 학생 데이터가 비어있습니다.")
            return False
        
        # 새 Excel 파일 생성
        wb = Workbook()
        ws = wb.active
        
        # 헤더 추가
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 학생 데이터 추가
        for row, student in enumerate(students, 2):
            for col, header in enumerate(headers, 1):
                value = student.get(header, '')
                ws.cell(row=row, column=col, value=value)
        
        # 파일 저장
        wb.save(excel_file)
        
        backup_date = backup_data.get('backup_date')
        print(f"✅ Firebase에서 학생 데이터 {len(students)}명 복원 완료")
        print(f"📅 백업 날짜: {backup_date}")
        
        return True
        
    except Exception as e:
        print(f"❌ 학생 데이터 복원 실패: {e}")
        return False

def auto_backup_and_restore():
    """
    자동 백업 및 복원 시스템
    1. students.xlsx가 있으면 Firebase에 백업
    2. students.xlsx가 없으면 Firebase에서 복원
    """
    db = init_firebase()
    if not db:
        return False
    
    excel_file = 'students.xlsx'
    
    if os.path.exists(excel_file):
        print("📁 students.xlsx 파일 발견 - Firebase 백업 시작")
        backup_success = backup_students_to_firebase(db, excel_file)
        if backup_success:
            print("✅ 자동 백업 완료")
        return backup_success
    else:
        print("📁 students.xlsx 파일 없음 - Firebase에서 복원 시작")
        restore_success = restore_students_from_firebase(db, excel_file)
        if restore_success:
            print("✅ 자동 복원 완료")
        return restore_success

if __name__ == "__main__":
    print("=== Replit 데이터 영구 보존 시스템 ===")
    auto_backup_and_restore()